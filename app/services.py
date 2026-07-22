import base64
import io
import re
from datetime import datetime
from email.message import EmailMessage
import mimetypes
import smtplib
from pathlib import Path
from typing import Iterable
from openpyxl import Workbook
from openpyxl.drawing.image import Image as XLImage
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from openpyxl.utils import get_column_letter
from PIL import Image, ImageOps
from .config import settings
from .schemas import MeerwerkInput

GENERATED = Path("data/generated")
UPLOADS = Path("data/uploads")
GENERATED.mkdir(parents=True, exist_ok=True)
UPLOADS.mkdir(parents=True, exist_ok=True)
THIN = Side(style="thin", color="B7B7B7")
BORDER = Border(left=THIN, right=THIN, top=THIN, bottom=THIN)
HEADER = PatternFill("solid", fgColor="EAF3EA")
GREEN = "218C42"


def safe_name(value: str) -> str:
    value = re.sub(r"[^A-Za-z0-9_-]+", "_", value.strip())
    return value[:60] or "meerwerk"


def decode_signature(data_url: str) -> bytes:
    if "," not in data_url:
        raise ValueError("Ongeldige handtekening")
    return base64.b64decode(data_url.split(",", 1)[1])


def normalize_photo(raw: bytes, destination: Path) -> None:
    image = Image.open(io.BytesIO(raw))
    image = ImageOps.exif_transpose(image).convert("RGB")
    image.thumbnail((1800, 1800))
    image.save(destination, "JPEG", quality=86, optimize=True)


def create_excel(data: MeerwerkInput, photo_paths: list[Path]) -> Path:
    wb = Workbook()
    ws = wb.active
    ws.title = "Meerwerk"
    ws.sheet_view.showGridLines = False
    ws.page_setup.orientation = "portrait"
    ws.page_setup.fitToWidth = 1
    ws.page_margins.left = .3
    ws.page_margins.right = .3
    for idx, width in enumerate([21, 18, 18, 18, 18, 18, 18, 18], 1):
        ws.column_dimensions[get_column_letter(idx)].width = width

    logo = Path("app/static/images/jan-bos-logo.png")
    if logo.exists():
        img = XLImage(str(logo)); img.width = 245; img.height = 100; ws.add_image(img, "A1")
    ws.merge_cells("E1:H2"); ws["E1"] = "Meerwerkrapport"
    ws["E1"].font = Font(size=23, bold=True, color=GREEN)
    ws["E1"].alignment = Alignment(horizontal="center", vertical="center")
    ws.merge_cells("E3:H3"); ws["E3"] = "Jan Bos Glas en Schilderwerken"
    ws["E3"].alignment = Alignment(horizontal="center")
    ws.merge_cells("E4:H4"); ws["E4"] = "024 - 388 6392  |  info@janbosschilderwerken.nl"
    ws["E4"].alignment = Alignment(horizontal="center")
    ws.row_dimensions[1].height = 34; ws.row_dimensions[2].height = 34

    fields = [
        ("Opdrachtgever", data.opdrachtgever, "Datum", data.datum),
        ("Opdrachtnummer", data.opdrachtnummer, "Medewerker", data.medewerker),
        ("Object / werkadres", data.object, "Aantal uren", data.uren),
    ]
    row = 6
    for l1, v1, l2, v2 in fields:
        ws[f"A{row}"] = l1; ws[f"A{row}"].font = Font(bold=True)
        ws.merge_cells(start_row=row, start_column=2, end_row=row, end_column=4); ws[f"B{row}"] = v1
        ws[f"E{row}"] = l2; ws[f"E{row}"].font = Font(bold=True)
        ws.merge_cells(start_row=row, start_column=6, end_row=row, end_column=8); ws[f"F{row}"] = v2
        for col in range(1, 9): ws.cell(row, col).border = BORDER
        row += 1

    sections = [
        ("Omschrijving van het meerwerk", data.omschrijving, 5),
        ("Reden / oorzaak", data.reden or "-", 3),
        ("Gebruikte materialen", data.materialen or "-", 4),
        ("Kostenindicatie", data.kostenindicatie or "Niet ingevuld", 1),
    ]
    row += 1
    for title, text, height in sections:
        ws.merge_cells(start_row=row, start_column=1, end_row=row, end_column=8)
        ws.cell(row, 1).value = title; ws.cell(row, 1).font = Font(bold=True, color="FFFFFF")
        ws.cell(row, 1).fill = PatternFill("solid", fgColor=GREEN)
        row += 1
        ws.merge_cells(start_row=row, start_column=1, end_row=row + height - 1, end_column=8)
        ws.cell(row, 1).value = text; ws.cell(row, 1).alignment = Alignment(wrap_text=True, vertical="top")
        for rr in range(row, row + height):
            for cc in range(1, 9): ws.cell(rr, cc).border = BORDER
        row += height + 1

    ws.merge_cells(start_row=row, start_column=1, end_row=row, end_column=8)
    ws.cell(row, 1).value = f"Foto's ({len(photo_paths)})"
    ws.cell(row, 1).font = Font(bold=True, color="FFFFFF"); ws.cell(row, 1).fill = PatternFill("solid", fgColor=GREEN)
    row += 1
    if photo_paths:
        for idx, photo in enumerate(photo_paths, 1):
            ws.merge_cells(start_row=row, start_column=1, end_row=row, end_column=8)
            ws.cell(row, 1).value = f"Foto {idx}"
            ws.cell(row, 1).font = Font(bold=True); ws.cell(row, 1).fill = HEADER
            row += 1
            ws.merge_cells(start_row=row, start_column=1, end_row=row + 11, end_column=8)
            for rr in range(row, row + 12):
                for cc in range(1, 9): ws.cell(rr, cc).border = BORDER
            img = XLImage(str(photo)); img.width = 520; img.height = 300; ws.add_image(img, f"B{row}")
            row += 13
    else:
        ws.merge_cells(start_row=row, start_column=1, end_row=row + 1, end_column=8)
        ws.cell(row, 1).value = "Geen foto's toegevoegd"; ws.cell(row, 1).alignment = Alignment(horizontal="center", vertical="center")
        row += 3

    ws.merge_cells(start_row=row, start_column=1, end_row=row, end_column=8)
    ws.cell(row, 1).value = "Handtekening klant / opdrachtgever"
    ws.cell(row, 1).font = Font(bold=True, color="FFFFFF"); ws.cell(row, 1).fill = PatternFill("solid", fgColor=GREEN)
    row += 1
    ws.merge_cells(start_row=row, start_column=1, end_row=row + 4, end_column=8)
    for rr in range(row, row + 5):
        for cc in range(1, 9): ws.cell(rr, cc).border = BORDER
    sig_path = GENERATED / f"sig_{datetime.now().strftime('%Y%m%d%H%M%S%f')}.png"
    Image.open(io.BytesIO(decode_signature(data.handtekening))).convert("RGBA").save(sig_path)
    sig = XLImage(str(sig_path)); sig.width = 280; sig.height = 90; ws.add_image(sig, f"C{row}")

    filename = f"Meerwerk_{safe_name(data.opdrachtgever)}_{safe_name(data.datum)}_{datetime.now().strftime('%Y%m%d_%H%M')}.xlsx"
    path = GENERATED / filename
    wb.save(path)
    sig_path.unlink(missing_ok=True)
    return path


def send_email(path: Path, data: MeerwerkInput, photo_paths: Iterable[Path]) -> bool:
    """Verstuur het Excel-rapport en de foto's via dezelfde SMTP-methode als de Glaszetter-app."""
    if not settings.smtp_host:
        raise RuntimeError("SMTP_HOST ontbreekt in Render Environment")
    recipient = settings.default_export_email.strip()
    if not recipient:
        raise RuntimeError("DEFAULT_EXPORT_EMAIL ontbreekt in Render Environment")

    photos = list(photo_paths)
    msg = EmailMessage()
    msg["Subject"] = f"Meerwerk - {data.opdrachtgever} - {data.object}"
    sender = settings.smtp_username.strip()
    if not sender:
        raise RuntimeError("SMTP_USERNAME ontbreekt in Render Environment")

    msg["From"] = sender
    msg["To"] = recipient
    msg.set_content(
        "Nieuw meerwerkrapport.\n\n"
        f"Medewerker: {data.medewerker}\n"
        f"Opdrachtgever: {data.opdrachtgever}\n"
        f"Object / werkadres: {data.object}\n"
        f"Datum: {data.datum}\n"
        f"Extra uren: {data.uren:g}\n"
        f"Aantal foto's: {len(photos)}\n\n"
        "Het Excel-rapport en de foto's zijn bijgevoegd."
    )

    msg.add_attachment(
        path.read_bytes(),
        maintype="application",
        subtype="vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        filename=path.name,
    )

    for index, photo in enumerate(photos, 1):
        mime, _ = mimetypes.guess_type(photo.name)
        maintype, subtype = (mime or "image/jpeg").split("/", 1)
        msg.add_attachment(
            photo.read_bytes(),
            maintype=maintype,
            subtype=subtype,
            filename=f"foto_{index}.jpg",
        )

    print(
        f"SMTP verzenden: van={sender} naar={recipient} "
        f"host={settings.smtp_host}:{settings.smtp_port} bijlagen={1 + len(photos)}",
        flush=True,
    )
    with smtplib.SMTP(settings.smtp_host, settings.smtp_port, timeout=45) as smtp:
        if settings.smtp_use_tls:
            smtp.starttls()
        if settings.smtp_username:
            smtp.login(settings.smtp_username, settings.smtp_password)
        smtp.send_message(msg)
    print("SMTP bevestigd: bericht geaccepteerd door mailserver", flush=True)
    return True
